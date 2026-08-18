import csv
import io
import re
from datetime import datetime
import pytz

BOGOTA_TZ = pytz.timezone('America/Bogota')

from flask import (
    Blueprint, render_template, request, jsonify,
    session, redirect, url_for, Response
)

from .db import get_db
from .auth import check_password, login_required

bp = Blueprint('main', __name__)

EMAIL_RE = re.compile(r'^[^\s@]+@[^\s@]+\.[^\s@]{2,}$')

# Rate limiting en memoria: {telefono: ultimo_timestamp}
_rate_limit_cache = {}
RATE_LIMIT_SECONDS = 5  # 1 envio cada 5 segundos por telefono


# ------------------------------------------------------------------
# Landing pública: una por cada QR (ref_code)
# ------------------------------------------------------------------
@bp.route('/producto/<ref_code>')
def producto(ref_code):
    db = get_db()
    cur = db.cursor()
    cur.execute(
        "SELECT * FROM productos WHERE ref_code = %s AND activo = 1",
        (ref_code,)
    )
    prod = cur.fetchone()

    if not prod:
        return render_template('404.html'), 404

    # Registrar el escaneo (mide interés real, aunque no deje datos)
    cur.execute("INSERT INTO escaneos (producto_id) VALUES (%s)", (prod['id'],))
    db.commit()

    cur.execute(
        "SELECT * FROM producto_media WHERE producto_id = %s ORDER BY orden",
        (prod['id'],)
    )
    media = cur.fetchall()

    specs = []
    if prod.get('especificaciones'):
        for parte in prod['especificaciones'].split('|'):
            if ':' in parte:
                k, v = parte.split(':', 1)
                specs.append((k.strip(), v.strip()))

    return render_template('producto.html', producto=prod, media=media, specs=specs)


# ------------------------------------------------------------------
# Home: si alguien entra sin ref_code, no debe ver un error feo
# ------------------------------------------------------------------
@bp.route('/')
def home():
    db = get_db()
    cur = db.cursor()


    cur.execute("SELECT ref_code, nombre, tipo_vidrio, descripcion, imagen_principal FROM productos WHERE activo = 1 ORDER BY nombre")
    vidrios = cur.fetchall()

    return render_template('home.html', vidrios=vidrios)

# ------------------------------------------------------------------
# Catálogo: 6 categorías
# ------------------------------------------------------------------
CATEGORIAS = [
    {'slug': 'design',         'nombre': 'Design',          'imagen': 'https://vidplex.com/wp-content/uploads/2023/02/design_.png'},
    {'slug': 'confort',        'nombre': 'Confort',         'imagen': 'https://vidplex.com/wp-content/uploads/2023/02/3d-house-interior.jpg'},
    {'slug': 'arquitectonico', 'nombre': 'Arquitectónico',  'imagen': 'https://vidplex.com/wp-content/uploads/2023/02/design_.png'},
    {'slug': 'control-solar',  'nombre': 'Control Solar',   'imagen': 'https://vidplex.com/wp-content/uploads/2023/02/controlsolar.png'},
    {'slug': 'seguridad',      'nombre': 'Seguridad',       'imagen': 'https://vidplex.com/wp-content/uploads/2023/02/seg.png'},
    {'slug': 'alto-desempeno', 'nombre': 'Alto Desempeño',  'imagen': 'https://vidplex.com/wp-content/uploads/2023/02/altode.png'},
]
@bp.route('/catalogo')
def catalogo():
    return render_template('catalogo.html', categorias=CATEGORIAS)


@bp.route('/catalogo/<categoria>')
def catalogo_categoria(categoria):
    cat = next((c for c in CATEGORIAS if c['slug'] == categoria), None)
    if not cat:
        return render_template('404.html'), 404

    db = get_db()
    cur = db.cursor()
    cur.execute(
        "SELECT ref_code, nombre, tipo_vidrio, descripcion, imagen_principal "
        "FROM productos WHERE activo = 1 AND categoria = %s ORDER BY nombre",
        (categoria,)
    )
    vidrios = list(cur.fetchall())

    return render_template('catalogo_categoria.html', vidrios=vidrios, categoria=cat)




# ------------------------------------------------------------------
# Captura de leads (llamado por fetch desde el formulario)
# ------------------------------------------------------------------
@bp.route('/leads', methods=['POST'])
def crear_lead():
    data = request.get_json(silent=True) or request.form

    nombre = (data.get('nombre') or '').strip()
    telefono = (data.get('telefono') or '').strip()
    correo = (data.get('correo') or '').strip().lower()
    tipo_proyecto = data.get('tipo_proyecto') or None
    autorizo = str(data.get('autorizo_datos', '')).lower() in ('1', 'true', 'on', 'si')
    ref_code = (data.get('ref_code') or '').strip()
    
    # HONEYPOT: campo invisible para humanos, si se llena = bot
    honeypot = (data.get('website') or '').strip()
    if honeypot:
        # Rechazo silencioso: el bot no sabe que fallo
        return jsonify({'ok': True}), 200

        # Rate limiting por telefono (anti doble-clic, no bloquea usuarios reales)
    ahora = datetime.now(BOGOTA_TZ).timestamp()
    ultimo_envio = _rate_limit_cache.get(telefono, 0)
    if ahora - ultimo_envio < RATE_LIMIT_SECONDS:
        return jsonify({'ok': False, 'error': 'Por favor espera unos segundos antes de reenviar.'}), 429
    _rate_limit_cache[telefono] = ahora

    # Validación en servidor — nunca confiar solo en el frontend
    errores = {}
    if len(nombre) < 4 or len(nombre.split()) < 2:
        errores['nombre'] = 'Escribe nombre y apellido.'

    tel_digits = ''.join(ch for ch in telefono if ch.isdigit())
    if not (7 <= len(tel_digits) <= 15):
        errores['telefono'] = 'Número de teléfono inválido.'

    if not EMAIL_RE.match(correo):
        errores['correo'] = 'Correo electrónico inválido.'

    if tipo_proyecto not in (None, '', 'residencial', 'comercial', 'constructora'):
        tipo_proyecto = None

    if not autorizo:
        errores['autorizo_datos'] = 'Debes autorizar el tratamiento de tus datos.'

    if errores:
        return jsonify({'ok': False, 'errores': errores}), 400

    db = get_db()
    cur = db.cursor()

    prod = None
    if ref_code:
        cur.execute("SELECT id FROM productos WHERE ref_code = %s", (ref_code,))
        prod = cur.fetchone()
        if not prod:
            return jsonify({'ok': False, 'error': 'Producto no encontrado.'}), 404

    # Deduplicar por TELEFONO (identificador unico de persona)
    # Si ya existe: NO se sobreescriben datos originales, solo se agrega vidrio visto
    cur.execute("SELECT id, nombre, correo FROM leads WHERE telefono = %s", (telefono,))
    existente = cur.fetchone()

    if existente:
        lead_id = existente['id']
        # NO actualizamos nombre ni correo — se conservan los originales
        if tipo_proyecto:
            cur.execute(
                "UPDATE leads SET tipo_proyecto = %s WHERE id = %s",
                (tipo_proyecto, lead_id)
            )
    else:
        cur.execute(
            """INSERT INTO leads
               (nombre, telefono, correo, tipo_proyecto, autorizo_datos, fecha_autorizacion)
               VALUES (%s, %s, %s, %s, 1, %s)""",
            (nombre, telefono, correo, tipo_proyecto, datetime.now(BOGOTA_TZ))
        )
        lead_id = cur.lastrowid

    # Vincular este lead con el vidrio que vio (sin duplicar si ya estaba)
    if prod:
        cur.execute(
            "INSERT IGNORE INTO lead_producto (lead_id, producto_id) VALUES (%s, %s)",
            (lead_id, prod['id'])
        )
    db.commit()

    return jsonify({'ok': True})


# ------------------------------------------------------------------
# Panel de administración
# ------------------------------------------------------------------
@bp.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    error = None
    if request.method == 'POST':
        correo = (request.form.get('correo') or '').strip().lower()
        password = request.form.get('password') or ''

        db = get_db()
        cur = db.cursor()
        cur.execute("SELECT * FROM admin_usuarios WHERE correo = %s", (correo,))
        admin = cur.fetchone()

        if admin and check_password(password, admin['password_hash']):
            session.clear()
            session['admin_id'] = admin['id']
            session['admin_usuario'] = admin['correo']
            return redirect(url_for('main.admin_dashboard'))

        error = 'Usuario o contraseña incorrectos.'

    return render_template('admin_login.html', error=error)


@bp.route('/admin/logout')
def admin_logout():
    session.clear()
    return redirect(url_for('main.admin_login'))


@bp.route('/admin')
@login_required
def admin_dashboard():
    return render_template('admin_dashboard.html', usuario=session.get('admin_usuario'))


@bp.route('/admin/reset-datos', methods=['POST'])
@login_required
def admin_reset_datos():
    """Borra TODOS los leads, sus vínculos con productos y el contador de escaneos.
    Acción irreversible — pensada para reiniciar el panel entre eventos/ferias."""
    db = get_db()
    cur = db.cursor()
    cur.execute("DELETE FROM lead_producto")
    cur.execute("DELETE FROM leads")
    cur.execute("DELETE FROM escaneos")
    db.commit()
    return jsonify({'ok': True})


@bp.route('/admin/data')
@login_required
def admin_data():
    """JSON que alimenta la tabla y la gráfica; el panel lo re-consulta solo."""
    db = get_db()
    cur = db.cursor()

    cur.execute("""
        SELECT p.ref_code, p.nombre,
               COUNT(DISTINCT e.id) AS escaneos,
               COUNT(DISTINCT lp.lead_id) AS leads
        FROM productos p
        LEFT JOIN escaneos e ON e.producto_id = p.id
        LEFT JOIN lead_producto lp ON lp.producto_id = p.id
        WHERE p.activo = 1
        GROUP BY p.id
        ORDER BY leads DESC, escaneos DESC
    """)
    ranking = cur.fetchall()

    cur.execute("""
        SELECT l.id, l.nombre, l.telefono, l.correo, l.tipo_proyecto, l.fecha_creacion,
               GROUP_CONCAT(p.ref_code ORDER BY p.ref_code SEPARATOR ', ') AS productos
        FROM leads l
        LEFT JOIN lead_producto lp ON lp.lead_id = l.id
        LEFT JOIN productos p ON p.id = lp.producto_id
        GROUP BY l.id
        ORDER BY l.fecha_creacion DESC
    """)
    leads = cur.fetchall()
    for row in leads:
        row['fecha_creacion'] = row['fecha_creacion'].strftime('%Y-%m-%d %H:%M')

    return jsonify({'ranking': ranking, 'leads': leads, 'total_leads': len(leads)})


@bp.route('/admin/export.xlsx')
@login_required
def admin_export():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    db = get_db()
    cur = db.cursor()
    cur.execute("""
        SELECT l.nombre, l.telefono, l.correo, l.tipo_proyecto, l.fecha_creacion,
               GROUP_CONCAT(p.ref_code ORDER BY p.ref_code SEPARATOR ', ') AS productos
        FROM leads l
        LEFT JOIN lead_producto lp ON lp.lead_id = l.id
        LEFT JOIN productos p ON p.id = lp.producto_id
        GROUP BY l.id
        ORDER BY l.fecha_creacion DESC
    """)
    rows = cur.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads VidPlex"
    
    # Encabezados con estilo
    headers = ['Nombre', 'Telefono', 'Correo', 'Tipo de proyecto', 'Fecha', 'Vidrios de interes']
    header_fill = PatternFill(start_color="C41E3A", end_color="C41E3A", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Datos
    for row_idx, r in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=r['nombre'])
        ws.cell(row=row_idx, column=2, value=r['telefono'])
        ws.cell(row=row_idx, column=3, value=r['correo'])
        ws.cell(row=row_idx, column=4, value=r['tipo_proyecto'] or '')
        ws.cell(row=row_idx, column=5, value=r['fecha_creacion'].strftime('%Y-%m-%d %H:%M') if r['fecha_creacion'] else '')
        ws.cell(row=row_idx, column=6, value=r['productos'] or '')
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 35
    
    # Guardar en buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"vidplex_leads_{datetime.now(BOGOTA_TZ).strftime('%Y%m%d_%H%M')}.xlsx"
    resp = Response(buffer.getvalue(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp.headers['Content-Disposition'] = f'attachment; filename={filename}'
    return resp
    filename = f"vidplex_leads_{datetime.now(BOGOTA_TZ).strftime('%Y%m%d_%H%M')}.xlsx"
    resp = Response(buffer.getvalue(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp.headers['Content-Disposition'] = f'attachment; filename={filename}'
    return resp
